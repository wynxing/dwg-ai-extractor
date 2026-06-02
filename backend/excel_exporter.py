from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import ExtractionResult, ExtractionTemplate


FAILURE_HEADERS = ["文件名", "文件路径", "错误信息", "耗时秒"]


def export_results_to_excel(results: list[ExtractionResult], template: ExtractionTemplate, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    result_sheet = workbook.active
    result_sheet.title = template.output_sheet_name or "提取结果"
    failure_sheet = workbook.create_sheet("失败清单")

    result_sheet.append([field.label for field in template.fields])
    for result in results:
        if result.status != "成功":
            continue
        result_sheet.append([result.values.get(field.key, "") for field in template.fields])

    failure_sheet.append(FAILURE_HEADERS)
    for result in results:
        if result.status != "成功":
            failure_sheet.append([result.file_path.name, str(result.file_path), result.error, round(result.elapsed_seconds, 3)])

    for sheet in [result_sheet, failure_sheet]:
        _style_sheet(sheet)

    workbook.save(output_path)


def _style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(color="000000", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        width = min(max(max_len + 2, 10), 48)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
