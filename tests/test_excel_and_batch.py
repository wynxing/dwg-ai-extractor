from pathlib import Path

from openpyxl import load_workbook

from backend.batch_processor import BatchProcessor
from backend.excel_exporter import export_results_to_excel
from backend.field_templates import DEFAULT_TEMPLATE
from backend.models import BatchJob, ExtractionResult, LLMConfig


def test_excel_export_uses_dynamic_template_columns(tmp_path: Path) -> None:
    output = tmp_path / "result.xlsx"
    results = [
        ExtractionResult(
            file_path=tmp_path / "a.dxf",
            status="成功",
            values={
                "sequence": "1",
                "item_code": "M1",
                "item_name": "零件",
                "image": "",
                "material": "PC",
                "color": "黑色",
                "surface_treatment": "喷油",
            },
        ),
        ExtractionResult(file_path=tmp_path / "bad.dxf", status="失败", error="bad"),
    ]

    export_results_to_excel(results, DEFAULT_TEMPLATE, output)

    workbook = load_workbook(output, data_only=True)
    assert workbook.sheetnames == ["物料明细表", "失败清单"]
    rows = list(workbook["物料明细表"].iter_rows(values_only=True))
    assert rows[0] == ("序号", "物料编码", "物料名称", "图片", "材料", "颜色", "表面处理")
    assert rows[1] == ("1", "M1", "零件", None, "PC", "黑色", "喷油")
    failures = list(workbook["失败清单"].iter_rows(values_only=True))
    assert len(failures) == 2


def test_batch_marks_file_failed_when_model_key_env_is_missing(tmp_path: Path) -> None:
    dxf = tmp_path / "a.dxf"
    dxf.write_text("not a real dxf", encoding="utf-8")
    output = tmp_path / "result.xlsx"
    job = BatchJob(
        input_paths=[dxf],
        output_excel_path=output,
        template=DEFAULT_TEMPLATE,
        llm=LLMConfig(model="fake", api_key_env_var="MISSING_ENV_FOR_TEST"),
    )

    results = BatchProcessor(job).run()

    assert results[0].status == "失败"
    assert output.exists()
